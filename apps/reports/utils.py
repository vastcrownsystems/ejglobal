# apps/reports/utils.py - Export and Chart Utilities

from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from datetime import datetime
import json


def generate_pdf_report(report, report_type):
    """
    Generate PDF report

    Args:
        report: Report model instance
        report_type: 'daily', 'monthly', or 'yearly'

    Returns:
        PDF content as bytes
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72,
                            topMargin=72, bottomMargin=18)

    # Container for elements
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0033A0'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#0033A0'),
        spaceAfter=12,
    )

    # Title
    if report_type == 'daily':
        title = f"Daily Sales Report - {report.report_date}"
    elif report_type == 'monthly':
        title = f"Monthly Sales Report - {report.month_name} {report.year}"
    elif report_type == 'yearly':
        title = f"Yearly Sales Report - {report.year}"

    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))

    # Company Info
    company_info = [
        ["EJ Global Services POS"],
        [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
    ]
    company_table = Table(company_info, colWidths=[6 * inch])
    company_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey),
    ]))
    elements.append(company_table)
    elements.append(Spacer(1, 20))

    # Sales Summary
    elements.append(Paragraph("Sales Summary", heading_style))

    summary_data = [
        ['Metric', 'Value'],
        ['Total Sales', f'₦{report.total_sales:,.2f}'],
        ['Total Orders', f'{report.total_orders:,}'],
        ['Items Sold', f'{report.total_items_sold:,}'],
        ['Average Order Value', f'₦{report.average_order_value:,.2f}'],
    ]

    summary_table = Table(summary_data, colWidths=[3 * inch, 3 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0033A0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Revenue Breakdown
    elements.append(Paragraph("Revenue Breakdown", heading_style))

    revenue_data = [
        ['Item', 'Amount'],
        ['Gross Sales', f'₦{report.gross_sales:,.2f}'],
        ['Discounts', f'-₦{report.total_discounts:,.2f}'],
        ['Tax', f'₦{report.total_tax:,.2f}'],
        ['Net Sales', f'₦{report.net_sales:,.2f}'],
    ]

    revenue_table = Table(revenue_data, colWidths=[3 * inch, 3 * inch])
    revenue_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7FBA00')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(revenue_table)
    elements.append(Spacer(1, 20))

    # Payment Methods
    elements.append(Paragraph("Payment Methods", heading_style))

    payment_data = [
        ['Method', 'Amount', 'Percentage'],
        ['Cash', f'₦{report.cash_sales:,.2f}',
         f'{(report.cash_sales / report.total_sales * 100) if report.total_sales > 0 else 0:.1f}%'],
        ['Card', f'₦{report.card_sales:,.2f}',
         f'{(report.card_sales / report.total_sales * 100) if report.total_sales > 0 else 0:.1f}%'],
        ['Transfer', f'₦{report.transfer_sales:,.2f}',
         f'{(report.transfer_sales / report.total_sales * 100) if report.total_sales > 0 else 0:.1f}%'],
        ['Mobile', f'₦{report.mobile_sales:,.2f}',
         f'{(report.mobile_sales / report.total_sales * 100) if report.total_sales > 0 else 0:.1f}%'],
    ]

    payment_table = Table(payment_data, colWidths=[2 * inch, 2 * inch, 2 * inch])
    payment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0033A0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(payment_table)

    # Build PDF
    doc.build(elements)

    # Get PDF content
    pdf = buffer.getvalue()
    buffer.close()

    return pdf


def generate_excel_report(report, report_type):
    """
    Generate Excel report with charts

    Args:
        report: Report model instance
        report_type: 'daily', 'monthly', or 'yearly'

    Returns:
        Excel content as bytes
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Styles
    header_fill = PatternFill(start_color="0033A0", end_color="0033A0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=16, color="0033A0")
    currency_format = '₦#,##0.00'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = "EJ Global POS - Sales Report"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')

    if report_type == 'daily':
        ws['A2'] = f"Daily Report - {report.report_date}"
    elif report_type == 'monthly':
        ws['A2'] = f"Monthly Report - {report.month_name} {report.year}"
    elif report_type == 'yearly':
        ws['A2'] = f"Yearly Report - {report.year}"

    ws.merge_cells('A2:D2')

    ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells('A3:D3')

    # Sales Summary
    row = 5
    ws[f'A{row}'] = "SALES SUMMARY"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')

    row += 1
    summary_data = [
        ['Total Sales', report.total_sales],
        ['Total Orders', report.total_orders],
        ['Items Sold', report.total_items_sold],
        ['Average Order Value', report.average_order_value],
    ]

    for label, value in summary_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        if isinstance(value, (int, float)) and value > 100:
            ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1

    # Revenue Breakdown
    row += 2
    ws[f'A{row}'] = "REVENUE BREAKDOWN"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')

    row += 1
    revenue_data = [
        ['Gross Sales', report.gross_sales],
        ['Discounts', report.total_discounts],
        ['Tax', report.total_tax],
        ['Net Sales', report.net_sales],
    ]

    for label, value in revenue_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1

    # Payment Methods
    row += 2
    ws[f'A{row}'] = "PAYMENT METHODS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:C{row}')

    row += 1
    ws[f'A{row}'] = "Method"
    ws[f'B{row}'] = "Amount"
    ws[f'C{row}'] = "Percentage"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row += 1
    payment_start_row = row
    payment_data = [
        ['Cash', report.cash_sales],
        ['Card', report.card_sales],
        ['Transfer', report.transfer_sales],
        ['Mobile', report.mobile_sales],
    ]

    for method, amount in payment_data:
        ws[f'A{row}'] = method
        ws[f'B{row}'] = amount
        ws[f'B{row}'].number_format = currency_format
        percentage = (amount / report.total_sales * 100) if report.total_sales > 0 else 0
        ws[f'C{row}'] = percentage / 100
        ws[f'C{row}'].number_format = '0.0%'
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = border
        row += 1

    # Add Pie Chart for Payment Methods
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=payment_start_row, max_row=payment_start_row + 3)
    data = Reference(ws, min_col=2, min_row=payment_start_row - 1, max_row=payment_start_row + 3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Payment Methods Distribution"
    ws.add_chart(pie, f"E{payment_start_row}")

    # Customer Metrics (if daily report)
    if report_type == 'daily' and hasattr(report, 'total_customers'):
        row += 2
        ws[f'A{row}'] = "CUSTOMER METRICS"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        customer_data = [
            ['Total Customers', report.total_customers],
            ['New Customers', report.new_customers],
            ['Returning Customers', report.returning_customers],
            ['Walk-in Sales', report.walk_in_sales],
        ]

        for label, value in customer_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    excel_content = buffer.getvalue()
    buffer.close()

    return excel_content


def prepare_chart_data(report, report_type):
    """
    Prepare data for Chart.js

    Args:
        report: Report model instance
        report_type: 'daily', 'monthly', or 'yearly'

    Returns:
        dict with chart data
    """
    chart_data = {}

    if report_type == 'daily':
        # Hourly sales data
        hourly_data = report.report_data.get('hourly_sales', [])
        chart_data['hourly'] = {
            'labels': [h['time'] for h in hourly_data],
            'sales': [h['sales'] for h in hourly_data],
            'orders': [h['orders'] for h in hourly_data],
        }

    elif report_type == 'monthly':
        # Daily breakdown
        daily_data = report.report_data.get('daily_breakdown', [])
        chart_data['daily'] = {
            'labels': [d['day'] for d in daily_data],
            'sales': [d['sales'] for d in daily_data],
            'orders': [d['orders'] for d in daily_data],
        }

    elif report_type == 'yearly':
        # Monthly breakdown
        monthly_data = report.report_data.get('monthly_breakdown', [])
        chart_data['monthly'] = {
            'labels': [m['month_name'] for m in monthly_data],
            'sales': [m['sales'] for m in monthly_data],
            'orders': [m['orders'] for m in monthly_data],
        }

    # Payment methods (for all types)
    chart_data['payment_methods'] = {
        'labels': ['Cash', 'Card', 'Transfer', 'Mobile'],
        'data': [
            float(report.cash_sales),
            float(report.card_sales),
            float(report.transfer_sales),
            float(report.mobile_sales),
        ]
    }

    return chart_data