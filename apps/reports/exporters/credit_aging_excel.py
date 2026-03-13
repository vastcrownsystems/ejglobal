from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class CreditAgingExcelExporter:

    def export(self, data, filepath):

        wb = Workbook()
        ws = wb.active
        ws.title = "Credit Aging"

        title_fill = PatternFill("solid", fgColor="0D47A1")
        header_fill = PatternFill("solid", fgColor="0D47A1")

        title_font = Font(color="FFFFFF", bold=True, size=16)
        header_font = Font(color="FFFFFF", bold=True)

        center = Alignment(horizontal="center")

        ws.merge_cells("A1:I1")

        title = ws["A1"]
        title.value = "CREDIT AGING REPORT"
        title.fill = title_fill
        title.font = title_font
        title.alignment = center

        ws.merge_cells("A2:I2")
        ws["A2"] = data["date_range"]
        ws["A2"].alignment = center

        row = 4

        headers = [
            "Customer",
            "Phone",
            "Current",
            "1-30 Days",
            "31-60 Days",
            "61-90 Days",
            "90+ Days",
            "Total"
        ]

        for col,h in enumerate(headers,1):

            cell = ws.cell(row=row,column=col)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        row += 1

        for c in data["customers"]:

            ws.cell(row=row,column=1,value=c["customer"])
            ws.cell(row=row,column=2,value=c["phone"])

            for i,key in enumerate(
                ["current","days_30","days_60","days_90","days_120","total"],3
            ):
                cell = ws.cell(row=row,column=i,value=c[key])
                cell.number_format='"₦"#,##0.00'

            row += 1

        widths=[26,18,14,14,14,14,14,16]

        for i,w in enumerate(widths,1):
            ws.column_dimensions[get_column_letter(i)].width=w

        wb.save(filepath)