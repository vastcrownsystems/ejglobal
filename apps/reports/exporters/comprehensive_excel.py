# apps/reports/exporters/comprehensive_excel.py
"""
Comprehensive Excel Report Exporter
Matches the Template.xlsx format:
  - Sheet 1: Inventory + Sales Details (side by side)
  - Sheet 2: Sales Summary
  - Sheet 3: Cashier Performance
  - Sheet 4: Customer Analysis
  - Sheet 5: Credit Aging
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from decimal import Decimal

from datetime import datetime

def _make_naive(dt):
    if isinstance(dt, datetime) and dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt


# ── Colour palette matching template ──────────────────────────────────────────
BLUE_DARK   = "0D47A1"   # dark blue header bg
WHITE       = "FFFFFF"
ALT_ROW     = "F8FAFC"   # light blue-grey alternate row
TOTAL_BG    = "E3F2FD"   # light blue totals row
ORANGE      = "FF8F00"   # accent for section titles
GREEN       = "1B5E20"
NAIRA_FMT   = '"₦"#,##0.00'
DATE_FMT    = "DD-MMM-YYYY"
CURRENCY    = NAIRA_FMT


def _hdr_font():
    return Font(name="Arial", bold=True, color=WHITE, size=10)

def _title_font(size=12):
    return Font(name="Arial", bold=True, size=size, color=BLUE_DARK)

def _body_font():
    return Font(name="Arial", size=9)

def _total_font():
    return Font(name="Arial", bold=True, size=9)

def _hdr_fill():
    return PatternFill("solid", fgColor=BLUE_DARK)

def _alt_fill():
    return PatternFill("solid", fgColor=ALT_ROW)

def _total_fill():
    return PatternFill("solid", fgColor=TOTAL_BG)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def _right():
    return Alignment(horizontal="right", vertical="center")

def _thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _col(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def _hdr_row(ws, row, cols, start_col=1):
    """Write a header row with dark blue background."""
    for i, label in enumerate(cols):
        c = ws.cell(row=row, column=start_col + i, value=label)
        c.font = _hdr_font()
        c.fill = _hdr_fill()
        c.alignment = _center()
        c.border = _thin_border()

def _data_row(ws, row, values, start_col=1, alt=False):
    """Write a data row with optional alternate shading."""
    fill = _alt_fill() if alt else None
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=val)
        c.font = _body_font()
        c.border = _thin_border()
        if fill:
            c.fill = fill
        if isinstance(val, Decimal):
            c.number_format = NAIRA_FMT
            c.alignment = _right()
        elif isinstance(val, (int, float)) and not isinstance(val, bool):
            c.alignment = _right()
        else:
            c.alignment = _left()

def _title_cell(ws, row, col, text, span_end=None, size=12):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _title_font(size)
    if span_end:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=span_end
        )
    return c

def _sub_label(ws, row, col, label, value, currency=False):
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = Font(name="Arial", size=9, color="555555")
    vc = ws.cell(row=row, column=col + 1, value=value)
    vc.font = Font(name="Arial", bold=True, size=9)
    if currency:
        vc.number_format = NAIRA_FMT
    return vc


class ComprehensiveExcelExporter:
    """
    Generates a multi-sheet Excel workbook containing:
      Sheet 1 — Inventory Report + Sales Details (side by side, template format)
      Sheet 2 — Sales Summary
      Sheet 3 — Cashier Performance
      Sheet 4 — Customer Analysis
      Sheet 5 — Credit & Overdue
    """

    def __init__(self, data: dict):
        self.data = data
        self.wb = Workbook()
        self.wb.remove(self.wb.active)   # remove default sheet

    # ── Public entry ──────────────────────────────────────────────────────────
    def build(self) -> Workbook:
        self._sheet_inventory_sales()
        self._sheet_sales_summary()
        self._sheet_cashier()
        self._sheet_customers()
        self._sheet_credit()
        return self.wb

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1: Inventory (cols A-J) + Sales Details (cols L-W)
    # ══════════════════════════════════════════════════════════════════════════
    def _sheet_inventory_sales(self):
        ws = self.wb.create_sheet("Inventory & Sales")
        inv  = self.data.get("inventory", {})
        sales_rows = self.data.get("sales_detail_rows", [])
        period = self.data.get("period_label", "")

        # ── Column widths ──
        widths = {
            "A": 20, "B": 10, "C": 22, "D": 12,
            "E": 14, "F": 10, "G": 10, "H": 10, "I": 14,
            "J": 14, "K": 14, "L": 16,  # J=Actual Closing, K=Unit Price, L=Stock Value
            "M": 4,  "N": 4,             # two free spacer columns
            "O": 26, "P": 14, "Q": 16, "R": 22,
            "S": 14, "T": 16, "U": 20, "V": 16,
            "W": 12, "X": 10, "Y": 14, "Z": 14,
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # ── INVENTORY section title (cols A-I) ──
        _title_cell(ws, 1, 1, "INVENTORY REPORT", span_end=9, size=13)
        ws.cell(row=2, column=1, value=f"Period: {period}").font = Font(name="Arial", size=9, italic=True)

        # inventory summary labels (rows 3-9)
        inv_summary = [
            ("Total Products",          inv.get("total_products", 0)),
            ("Total Opening Stock",     inv.get("total_opening_stock", 0)),
            ("Total Stock Added",       inv.get("total_stock_added", 0)),
            ("Adjustments Increase",    inv.get("total_adjustments_increase", 0)),
            ("Adjustments Decrease",    inv.get("total_adjustments_decrease", 0)),
            ("Total Quantity Sold",     inv.get("total_quantity_sold", 0)),
            ("Expected Closing",        inv.get("total_expected_closing", 0)),
            ("Actual Closing",          inv.get("total_actual_closing", 0)),
            ("Total Variance",          inv.get("total_variance", 0)),
            ("Total Stock Value",       inv.get("total_stock_value", Decimal("0.00"))),
        ]
        for i, (label, val) in enumerate(inv_summary):
            r = 3 + i
            lc = ws.cell(row=r, column=1, value=label)
            lc.font = Font(name="Arial", size=9)
            vc = ws.cell(row=r, column=2, value=val)
            vc.font = Font(name="Arial", bold=True, size=9)
            if isinstance(val, Decimal):
                vc.number_format = NAIRA_FMT
                vc.alignment = _right()

        # inventory product table header
        inv_hdr_row = 14
        inv_hdrs = ["Product Name", "Variant", "SKU", "Category",
                    "Opening Stock", "Adj (+)", "Adj (-)", "Sales",
                    "Expected Closing", "Actual Closing", "Unit Price (₦)", "Stock Value (₦)"]
        # Inventory table: cols A-L (12 columns)
        _hdr_row(ws, inv_hdr_row, inv_hdrs, start_col=1)

        products = inv.get("products", [])
        for idx, p in enumerate(products):
            r = inv_hdr_row + 1 + idx
            alt = idx % 2 == 1
            row_data = [
                p.get("product_name", ""),
                p.get("variant_name", ""),
                p.get("sku", ""),
                p.get("category", ""),
                p.get("opening_stock", 0),
                p.get("adjustments_increase", 0),
                p.get("adjustments_decrease", 0),
                p.get("sales", 0),
                p.get("expected_closing", 0),
                p.get("actual_closing", 0),
                p.get("unit_price", Decimal("0.00")),
                p.get("stock_value", Decimal("0.00")),
            ]
            _data_row(ws, r, row_data, start_col=1, alt=alt)

        # totals row
        if products:
            tr = inv_hdr_row + 1 + len(products)
            ws.cell(row=tr, column=1, value="TOTALS").font = _total_font()
            for col_idx, key in enumerate(
                ["", "", "", "",
                 "total_opening_stock", None, None,
                 "total_quantity_sold", "total_expected_closing",
                 "total_actual_closing", None, "total_stock_value"],
                start=1
            ):
                if key and key in inv:
                    c = ws.cell(row=tr, column=col_idx, value=inv[key])
                    c.font = _total_font()
                    c.fill = _total_fill()
                    if isinstance(inv[key], Decimal):
                        c.number_format = NAIRA_FMT

        # ── SALES DETAILS section (cols L-W) ──
        # L=12, M=13, N=14, O=15, P=16, Q=17, R=18, S=19, T=20, U=21, V=22, W=23
        SALES_START = 15   # column O (2 free spacer cols M & N)
        _title_cell(ws, 1, SALES_START, "Sales Details", span_end=SALES_START + 11, size=13)
        ws.cell(row=2, column=SALES_START, value=f"Period: {period}").font = Font(name="Arial", size=9, italic=True)

        sales_hdrs = [
            "Order Number", "Date", "Cashier", "Customer",
            "Customer Type", "Sales Person", "Product", "Variant",
            "Unit Price (₦)", "Quantity", "Sub-Total (₦)", "Payment Type"
        ]
        _hdr_row(ws, 2, sales_hdrs, start_col=SALES_START)

        for idx, row in enumerate(sales_rows):
            r = 3 + idx
            alt = idx % 2 == 1
            date_val = _make_naive(row.get("date"))
            row_data = [
                row.get("order_number", ""),
                date_val,
                row.get("cashier", ""),
                row.get("customer", "Walk-in"),
                row.get("customer_type", "Walk-in"),
                row.get("sales_person", "—"),
                row.get("product", ""),
                row.get("variant", ""),
                row.get("unit_price", Decimal("0.00")),
                row.get("quantity", 0),
                row.get("line_total", Decimal("0.00")),
                row.get("payment_type", ""),
            ]
            for i, val in enumerate(row_data):
                c = ws.cell(row=r, column=SALES_START + i, value=val)
                c.font = _body_font()
                c.border = _thin_border()
                if alt:
                    c.fill = _alt_fill()
                if isinstance(val, Decimal):
                    c.number_format = NAIRA_FMT
                    c.alignment = _right()
                elif hasattr(val, "strftime"):
                    c.number_format = DATE_FMT
                    c.alignment = _center()
                elif isinstance(val, (int, float)) and not isinstance(val, bool):
                    c.alignment = _right()
                else:
                    c.alignment = _left()

        # totals row for sales
        if sales_rows:
            tr = 3 + len(sales_rows)
            ws.cell(row=tr, column=SALES_START, value="TOTALS").font = _total_font()
            # sub-total formula
            stcol = SALES_START + 10   # column V
            qty_col = SALES_START + 9  # column U
            first_data = 3
            last_data = 3 + len(sales_rows) - 1
            stcol_l = get_column_letter(stcol)
            qtycol_l = get_column_letter(qty_col)
            c_st = ws.cell(row=tr, column=stcol,
                           value=f"=SUM({stcol_l}{first_data}:{stcol_l}{last_data})")
            c_st.font = _total_font()
            c_st.fill = _total_fill()
            c_st.number_format = NAIRA_FMT
            c_qty = ws.cell(row=tr, column=qty_col,
                            value=f"=SUM({qtycol_l}{first_data}:{qtycol_l}{last_data})")
            c_qty.font = _total_font()
            c_qty.fill = _total_fill()

        ws.freeze_panes = None

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: Sales Summary
    # ══════════════════════════════════════════════════════════════════════════
    def _sheet_sales_summary(self):
        ws = self.wb.create_sheet("Sales Summary")
        ss   = self.data.get("sales_summary", {})
        period = self.data.get("period_label", "")

        for col, w in {"A": 30, "B": 20, "C": 2, "D": 30, "E": 20}.items():
            ws.column_dimensions[col].width = w

        _title_cell(ws, 1, 1, "SALES SUMMARY REPORT", span_end=5, size=13)
        ws.cell(row=2, column=1, value=f"Period: {period}").font = Font(name="Arial", size=9, italic=True)

        # Left block — key metrics
        left = [
            ("Total Orders",          ss.get("total_orders", 0),         False),
            ("Total Items Sold",      ss.get("total_items_sold", 0),      False),
            ("Total Revenue",         ss.get("total_sales", Decimal("0")), True),
            ("Gross Sales",           ss.get("gross_sales", Decimal("0")), True),
            ("Net Sales",             ss.get("net_sales", Decimal("0")),   True),
            ("Total Discounts",       ss.get("total_discounts", Decimal("0")), True),
            ("Average Order Value",   ss.get("average_order_value", Decimal("0")), True),
        ]
        for i, (lbl, val, cur) in enumerate(left):
            r = 4 + i
            ws.cell(row=r, column=1, value=lbl).font = Font(name="Arial", size=9)
            vc = ws.cell(row=r, column=2, value=val)
            vc.font = Font(name="Arial", bold=True, size=9)
            if cur:
                vc.number_format = NAIRA_FMT
                vc.alignment = _right()

        # Right block — payment breakdown
        _title_cell(ws, 4, 4, "PAYMENT BREAKDOWN", size=10)
        right = [
            ("Cash Sales",        ss.get("cash_sales", Decimal("0"))),
            ("Card Sales",        ss.get("card_sales", Decimal("0"))),
            ("Transfer Sales",    ss.get("transfer_sales", Decimal("0"))),
            ("Credit Sales",      ss.get("credit_sales", Decimal("0"))),
            ("Total Paid",        ss.get("total_sales", Decimal("0"))),
        ]
        for i, (lbl, val) in enumerate(right):
            r = 5 + i
            ws.cell(row=r, column=4, value=lbl).font = Font(name="Arial", size=9)
            vc = ws.cell(row=r, column=5, value=val)
            vc.font = Font(name="Arial", bold=True, size=9)
            vc.number_format = NAIRA_FMT
            vc.alignment = _right()

        # Customer metrics
        r_start = 14
        _title_cell(ws, r_start, 1, "CUSTOMER METRICS", size=10)
        cust = [
            ("Total Customers",     ss.get("total_customers", 0)),
            ("New Customers",       ss.get("new_customers", 0)),
            ("Returning Customers", ss.get("returning_customers", 0)),
            ("Walk-in Sales",       ss.get("walk_in_sales", 0)),
        ]
        for i, (lbl, val) in enumerate(cust):
            r = r_start + 1 + i
            ws.cell(row=r, column=1, value=lbl).font = Font(name="Arial", size=9)
            ws.cell(row=r, column=2, value=val).font = Font(name="Arial", bold=True, size=9)

        # Product performance table
        r_tbl = r_start + 8
        _title_cell(ws, r_tbl, 1, "TOP PRODUCTS", size=10)

        prod_hdrs = ["Product", "Variant", "SKU", "Qty Sold", "Revenue (₦)"]
        for col_w, col_l in zip([30, 14, 24, 12, 18], ["A", "B", "C", "D", "E"]):
            ws.column_dimensions[col_l].width = col_w
        _hdr_row(ws, r_tbl + 1, prod_hdrs)

        products = self.data.get("sales_summary", {}).get("products", [])
        for idx, p in enumerate(products):
            r = r_tbl + 2 + idx
            _data_row(ws, r, [
                p.get("product_name", ""),
                p.get("variant_name", ""),
                p.get("sku", ""),
                p.get("quantity_sold", 0),
                p.get("revenue", Decimal("0")),
            ], alt=idx % 2 == 1)

        # Daily trend table
        trend = self.data.get("daily_trend", [])
        if trend:
            r_trend = r_tbl + 2 + len(products) + 3
            _title_cell(ws, r_trend, 1, "DAILY SALES TREND", size=10)
            _hdr_row(ws, r_trend + 1, ["Date", "Orders", "Revenue (₦)", "Items Sold"])
            for idx, t in enumerate(trend):
                r = r_trend + 2 + idx
                _data_row(ws, r, [
                    t.get("day"),
                    t.get("total_orders", 0),
                    t.get("total_sales", Decimal("0")),
                    t.get("total_items", 0),
                ], alt=idx % 2 == 1)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3: Cashier Performance
    # ══════════════════════════════════════════════════════════════════════════
    def _sheet_cashier(self):
        ws = self.wb.create_sheet("Cashier Performance")
        data = self.data.get("cashier_data", {})
        period = self.data.get("period_label", "")

        for col, w in {"A": 22, "B": 12, "C": 12, "D": 18, "E": 14, "F": 20, "G": 20}.items():
            ws.column_dimensions[col].width = w

        _title_cell(ws, 1, 1, "CASHIER PERFORMANCE REPORT", span_end=7, size=13)
        ws.cell(row=2, column=1, value=f"Period: {period}").font = Font(name="Arial", size=9, italic=True)

        summary = data.get("summary", {})
        summ_labels = [
            ("Total Cashiers",  summary.get("total_cashiers", 0),  False),
            ("Total Orders",    summary.get("total_orders", 0),    False),
            ("Total Revenue",   summary.get("total_revenue", Decimal("0")), True),
            ("Total Sessions",  summary.get("total_sessions", 0),  False),
        ]
        for i, (lbl, val, cur) in enumerate(summ_labels):
            r = 4 + i
            ws.cell(row=r, column=1, value=lbl).font = Font(name="Arial", size=9)
            vc = ws.cell(row=r, column=2, value=val)
            vc.font = Font(name="Arial", bold=True, size=9)
            if cur:
                vc.number_format = NAIRA_FMT

        # Cashier table
        r_tbl = 10
        _title_cell(ws, r_tbl, 1, "PER-CASHIER BREAKDOWN", size=10)
        _hdr_row(ws, r_tbl + 1, ["Cashier", "Orders", "Items Sold", "Revenue (₦)", "Avg Order (₦)", "First Sale", "Last Sale"])
        cashiers = data.get("cashiers", [])
        for idx, c in enumerate(cashiers):
            r = r_tbl + 2 + idx
            _data_row(ws, r, [
                c.get("cashier", ""),
                c.get("orders", 0),
                c.get("items_sold", 0),
                c.get("revenue", Decimal("0")),
                c.get("avg_order", Decimal("0")),
                _make_naive(c.get("first_sale")),
                _make_naive(c.get("last_sale"))
            ], alt=idx % 2 == 1)

        # Sessions table
        sessions = data.get("sessions", [])
        if sessions:
            r_sess = r_tbl + 2 + len(cashiers) + 3
            _title_cell(ws, r_sess, 1, "SESSION BREAKDOWN", size=10)
            for col, w in {"A": 10, "B": 20, "C": 20, "D": 22,
                           "E": 20, "F": 20, "G": 10, "H": 18}.items():
                ws.column_dimensions[col].width = w
            _hdr_row(ws, r_sess + 1, ["Session ID", "Store", "Register", "Cashier",
                                       "Opened", "Closed", "Orders", "Revenue (₦)"])
            for idx, s in enumerate(sessions):
                r = r_sess + 2 + idx
                _data_row(ws, r, [
                    str(s.get("session", "")),
                    s.get("store", ""),
                    s.get("register", ""),
                    s.get("cashier", ""),
                    _make_naive(s.get("opened")),
                    _make_naive(s.get("closed")),
                    s.get("orders", 0),
                    s.get("revenue", Decimal("0")),
                ], alt=idx % 2 == 1)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4: Customer Analysis
    # ══════════════════════════════════════════════════════════════════════════
    def _sheet_customers(self):
        ws = self.wb.create_sheet("Customer Analysis")
        data = self.data.get("customer_data", {})
        period = self.data.get("period_label", "")

        for col, w in {"A": 24, "B": 14, "C": 16, "D": 12,
                       "E": 12, "F": 18, "G": 18, "H": 18, "I": 16}.items():
            ws.column_dimensions[col].width = w

        _title_cell(ws, 1, 1, "CUSTOMER ANALYSIS REPORT", span_end=9, size=13)
        ws.cell(row=2, column=1, value=f"Period: {period}").font = Font(name="Arial", size=9, italic=True)

        summary = data.get("summary", {})
        for i, (lbl, val, cur) in enumerate([
            ("Total Customers",       summary.get("total_customers", 0),              False),
            ("Total Revenue",         summary.get("total_revenue", Decimal("0")),     True),
            ("Avg Customer Value",    summary.get("avg_customer_value", Decimal("0")),True),
        ]):
            r = 4 + i
            ws.cell(row=r, column=1, value=lbl).font = Font(name="Arial", size=9)
            vc = ws.cell(row=r, column=2, value=val)
            vc.font = Font(name="Arial", bold=True, size=9)
            if cur:
                vc.number_format = NAIRA_FMT

        r_tbl = 9
        _title_cell(ws, r_tbl, 1, "CUSTOMER BREAKDOWN", size=10)
        _hdr_row(ws, r_tbl + 1, [
            "Customer", "Phone", "Customer Type", "Sales Person",
            "Orders", "Items", "Revenue (₦)", "Avg Order (₦)",
            "First Purchase", "Last Purchase"
        ])
        for col, w in {"G": 18, "H": 16, "I": 18, "J": 18}.items():
            ws.column_dimensions[col].width = w

        customers = data.get("customers", [])
        for idx, c in enumerate(customers):
            r = r_tbl + 2 + idx
            _data_row(ws, r, [
                c.get("customer", ""),
                c.get("phone", ""),
                c.get("customer_type", ""),
                c.get("sales_person", "—"),
                c.get("orders", 0),
                c.get("items", 0),
                c.get("revenue", Decimal("0")),
                c.get("avg_order", Decimal("0")),
                _make_naive(c.get("first_purchase")),
                _make_naive(c.get("first_purchase"))
            ], alt=idx % 2 == 1)

        # totals
        if customers:
            tr = r_tbl + 2 + len(customers)
            ws.cell(row=tr, column=1, value="TOTALS").font = _total_font()
            rev_col = get_column_letter(8)   # column H
            first_r = r_tbl + 2
            last_r  = tr - 1
            c = ws.cell(row=tr, column=8,
                        value=f"=SUM({rev_col}{first_r}:{rev_col}{last_r})")
            c.font = _total_font()
            c.fill = _total_fill()
            c.number_format = NAIRA_FMT

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5: Credit & Overdue
    # ══════════════════════════════════════════════════════════════════════════
    def _sheet_credit(self):
        ws = self.wb.create_sheet("Credit & Overdue")
        aging = self.data.get("credit_aging", {})
        overdue = self.data.get("overdue_credit", {})
        period = self.data.get("period_label", "")

        for col, w in {"A": 24, "B": 14, "C": 14, "D": 14,
                       "E": 14, "F": 14, "G": 16, "H": 14}.items():
            ws.column_dimensions[col].width = w

        # Credit Aging
        _title_cell(ws, 1, 1, "CREDIT AGING REPORT", span_end=7, size=13)
        ws.cell(row=2, column=1, value=aging.get("date_range", "")).font = Font(name="Arial", size=9, italic=True)

        aging_summ = aging.get("summary", {})
        for i, (lbl, val) in enumerate([
            ("Total Outstanding", aging_summ.get("total_outstanding", Decimal("0"))),
            ("Current",           aging_summ.get("current", Decimal("0"))),
            ("1-30 Days",         aging_summ.get("days_30", Decimal("0"))),
            ("31-60 Days",        aging_summ.get("days_60", Decimal("0"))),
            ("61-90 Days",        aging_summ.get("days_90", Decimal("0"))),
            ("90+ Days",          aging_summ.get("days_120", Decimal("0"))),
        ]):
            r = 4 + i
            ws.cell(row=r, column=1, value=lbl).font = Font(name="Arial", size=9)
            vc = ws.cell(row=r, column=2, value=val)
            vc.font = Font(name="Arial", bold=True, size=9)
            vc.number_format = NAIRA_FMT

        r_tbl = 12
        _title_cell(ws, r_tbl, 1, "AGING BY CUSTOMER", size=10)
        _hdr_row(ws, r_tbl + 1, [
            "Customer", "Phone", "Current", "1-30 Days",
            "31-60 Days", "61-90 Days", "90+ Days", "Total Outstanding"
        ])
        aging_customers = aging.get("customers", [])
        for idx, c in enumerate(aging_customers):
            r = r_tbl + 2 + idx
            _data_row(ws, r, [
                c.get("customer", ""),
                c.get("phone", ""),
                c.get("current", Decimal("0")),
                c.get("days_30", Decimal("0")),
                c.get("days_60", Decimal("0")),
                c.get("days_90", Decimal("0")),
                c.get("days_120", Decimal("0")),
                c.get("total", Decimal("0")),
            ], alt=idx % 2 == 1)

        # Overdue section
        r_over = r_tbl + 2 + len(aging_customers) + 4
        _title_cell(ws, r_over, 1, "OVERDUE CREDIT DETAIL", span_end=7, size=12)
        ws.cell(row=r_over + 1, column=1, value=overdue.get("date_range", "")).font = Font(name="Arial", size=9, italic=True)

        over_summ = overdue.get("summary", {})
        ws.cell(row=r_over + 2, column=1, value="Total Overdue Accounts").font = Font(name="Arial", size=9)
        ws.cell(row=r_over + 2, column=2, value=over_summ.get("accounts", 0)).font = Font(name="Arial", bold=True, size=9)
        ws.cell(row=r_over + 3, column=1, value="Total Overdue Amount").font = Font(name="Arial", size=9)
        vc = ws.cell(row=r_over + 3, column=2, value=over_summ.get("total_overdue", Decimal("0")))
        vc.font = Font(name="Arial", bold=True, size=9)
        vc.number_format = NAIRA_FMT

        r_otbl = r_over + 5
        _hdr_row(ws, r_otbl, [
            "Customer", "Phone", "Invoice #", "Due Date",
            "Days Overdue", "Outstanding (₦)", "Priority"
        ])
        for idx, row in enumerate(overdue.get("rows", [])):
            r = r_otbl + 1 + idx
            _data_row(ws, r, [
                row.get("customer", ""),
                row.get("phone", ""),
                row.get("invoice", ""),
                row.get("due_date"),
                row.get("days_overdue", 0),
                row.get("outstanding", Decimal("0")),
                row.get("status", ""),
            ], alt=idx % 2 == 1)