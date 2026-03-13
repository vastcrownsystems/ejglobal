from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class LowStockExcelExporter:

    def export(self, data, filepath):

        wb = Workbook()
        ws = wb.active
        ws.title = "Low Stock Alert"

        # =========================
        # STYLES
        # =========================

        title_fill = PatternFill("solid", fgColor="0D47A1")
        header_fill = PatternFill("solid", fgColor="0D47A1")

        status_low = PatternFill("solid", fgColor="FFF3CD")
        status_critical = PatternFill("solid", fgColor="FFE5B4")
        status_out = PatternFill("solid", fgColor="F8D7DA")

        title_font = Font(color="FFFFFF", bold=True, size=16)
        header_font = Font(color="FFFFFF", bold=True)

        center = Alignment(horizontal="center", vertical="center")

        # =========================
        # TITLE
        # =========================

        ws.merge_cells("A1:H1")
        cell = ws["A1"]
        cell.value = "LOW STOCK ALERT REPORT"
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = center

        ws.merge_cells("A2:H2")
        ws["A2"] = f"Period: {data.get('date_range','')}"
        ws["A2"].alignment = center

        # =========================
        # SUMMARY
        # =========================

        summary = data["summary"]

        ws["A4"] = "Total Products"
        ws["B4"] = summary["total_products"]

        ws["A5"] = "Low Stock Items"
        ws["B5"] = summary["low_stock_items"]

        ws["A6"] = "Out of Stock"
        ws["B6"] = summary["out_of_stock"]

        ws["A7"] = "Critical Level"
        ws["B7"] = summary["critical"]

        # =========================
        # TABLE HEADER
        # =========================

        headers = [
            "Product Name",
            "Variant",
            "SKU",
            "Category",
            "Current Stock",
            "Reorder Level",
            "Shortage",
            "Status"
        ]

        row = 10

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # =========================
        # DATA ROWS
        # =========================

        row += 1

        for item in data["products"]:

            ws.cell(row=row, column=1, value=item["product_name"])
            ws.cell(row=row, column=2, value=item["variant_name"])
            ws.cell(row=row, column=3, value=item["sku"])
            ws.cell(row=row, column=4, value=item["category"])
            ws.cell(row=row, column=5, value=item["stock_quantity"])
            ws.cell(row=row, column=6, value=item["reorder_level"])
            ws.cell(row=row, column=7, value=item["difference"])

            status_cell = ws.cell(row=row, column=8, value=item["status"])

            # STATUS COLORS
            if item["status"] == "OUT OF STOCK":
                status_cell.fill = status_out
            elif item["status"] == "CRITICAL":
                status_cell.fill = status_critical
            elif item["status"] == "LOW":
                status_cell.fill = status_low

            row += 1

        # =========================
        # COLUMN WIDTH
        # =========================

        widths = [28, 12, 28, 18, 15, 16, 12, 14]

        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(filepath)