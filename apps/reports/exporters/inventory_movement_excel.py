from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class InventoryMovementExcelExporter:

    def export(self, data, filepath):

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Movement"

        # ======================
        # STYLES
        # ======================

        title_fill = PatternFill("solid", fgColor="0D47A1")
        header_fill = PatternFill("solid", fgColor="0D47A1")
        alt_fill = PatternFill("solid", fgColor="F8FAFC")

        title_font = Font(color="FFFFFF", bold=True, size=16)
        header_font = Font(color="FFFFFF", bold=True)

        center = Alignment(horizontal="center", vertical="center")

        # ======================
        # TITLE
        # ======================

        ws.merge_cells("A1:H1")

        cell = ws["A1"]
        cell.value = "INVENTORY MOVEMENT REPORT"
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = center

        ws.merge_cells("A2:H2")
        ws["A2"] = f"Period: {data.get('date_range','')}"
        ws["A2"].alignment = center

        # ======================
        # SUMMARY
        # ======================

        summary = data.get("summary", {})

        ws["A4"] = "Total Movements"
        ws["B4"] = summary.get("total_movements", 0)

        # ======================
        # TABLE HEADER
        # ======================

        headers = [
            "Date",
            "Product",
            "Variant",
            "SKU",
            "Movement Type",
            "Quantity",
            "Stock Before",
            "Stock After"
        ]

        row = 7

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # ======================
        # DATA
        # ======================

        row += 1

        for i, m in enumerate(data.get("movements", [])):

            ws.cell(row=row, column=1, value=str(m["date"]))
            ws.cell(row=row, column=2, value=m["product_name"])
            ws.cell(row=row, column=3, value=m["variant_name"])
            ws.cell(row=row, column=4, value=m["sku"])
            ws.cell(row=row, column=5, value=m["movement_type"])
            ws.cell(row=row, column=6, value=m["quantity"])
            ws.cell(row=row, column=7, value=m["stock_before"])
            ws.cell(row=row, column=8, value=m["stock_after"])

            if i % 2 == 1:
                for c in range(1, 9):
                    ws.cell(row=row, column=c).fill = alt_fill

            row += 1

        # ======================
        # COLUMN WIDTH
        # ======================

        widths = [20, 28, 14, 28, 20, 12, 15, 15]

        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(filepath)