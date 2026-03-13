from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class DailyInventoryExcelExporter:

    def export(self, data, filepath):

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Inventory"

        # =========================
        # STYLES
        # =========================

        title_fill = PatternFill("solid", fgColor="0D47A1")
        header_fill = PatternFill("solid", fgColor="0D47A1")

        alt_row_fill = PatternFill("solid", fgColor="F8FAFC")

        variance_positive = PatternFill("solid", fgColor="FFE5B4")
        variance_negative = PatternFill("solid", fgColor="F8D7DA")

        title_font = Font(color="FFFFFF", bold=True, size=16)
        header_font = Font(color="FFFFFF", bold=True)

        center = Alignment(horizontal="center", vertical="center")

        # =========================
        # TITLE
        # =========================

        ws.merge_cells("A1:M1")

        title_cell = ws["A1"]
        title_cell.value = "DAILY INVENTORY REPORT"
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center

        ws.merge_cells("A2:M2")

        period_cell = ws["A2"]
        period_cell.value = f"Period: {data.get('date_range','')}"
        period_cell.alignment = center

        # =========================
        # SUMMARY
        # =========================

        summary = data.get("summary", {})

        ws["A4"] = "Total Products"
        ws["B4"] = summary.get("total_products", 0)

        ws["A5"] = "Total Opening Stock"
        ws["B5"] = summary.get("total_opening_stock", 0)

        ws["A6"] = "Total Stock Added"
        ws["B6"] = summary.get("total_stock_added", 0)

        ws["A7"] = "Adjustments Increase"
        ws["B7"] = summary.get("total_adjustments_increase", 0)

        ws["A8"] = "Adjustments Decrease"
        ws["B8"] = summary.get("total_adjustments_decrease", 0)

        ws["A9"] = "Total Quantity Sold"
        ws["B9"] = summary.get("total_quantity_sold", 0)

        ws["A10"] = "Expected Closing"
        ws["B10"] = summary.get("total_expected_closing", 0)

        ws["A11"] = "Actual Closing"
        ws["B11"] = summary.get("total_actual_closing", 0)

        ws["A12"] = "Total Variance"
        ws["B12"] = summary.get("total_variance", 0)

        ws["A13"] = "Total Stock Value"
        value_cell = ws["B13"]
        value_cell.value = summary.get("total_revenue", Decimal("0.00"))
        value_cell.number_format = '"₦"#,##0.00'

        # =========================
        # TABLE HEADER
        # =========================

        headers = [
            "Product Name",
            "Variant",
            "SKU",
            "Category",
            "Opening Stock",
            "Stock Added",
            "Adj (+)",
            "Adj (-)",
            "Sales",
            "Expected Closing",
            "Actual Closing",
            "Variance",
            "Stock Value",
        ]

        row = 16

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # =========================
        # DATA ROWS
        # =========================

        row += 1

        for index, item in enumerate(data.get("products", [])):

            ws.cell(row=row, column=1, value=item.get("product_name"))
            ws.cell(row=row, column=2, value=item.get("variant_name"))
            ws.cell(row=row, column=3, value=item.get("sku"))
            ws.cell(row=row, column=4, value=item.get("category"))
            ws.cell(row=row, column=5, value=item.get("opening_stock"))
            ws.cell(row=row, column=6, value=item.get("additions"))
            ws.cell(row=row, column=7, value=item.get("adjustments_increase"))
            ws.cell(row=row, column=8, value=item.get("adjustments_decrease"))
            ws.cell(row=row, column=9, value=item.get("sales"))
            ws.cell(row=row, column=10, value=item.get("expected_closing"))
            ws.cell(row=row, column=11, value=item.get("actual_closing"))

            variance_cell = ws.cell(row=row, column=12, value=item.get("variance"))

            value_cell = ws.cell(row=row, column=13, value=item.get("stock_value"))

            value_cell.number_format = '"₦"#,##0.00'

            # =========================
            # ALTERNATE ROW COLOR
            # =========================

            if index % 2 == 1:
                for c in range(1, 14):
                    ws.cell(row=row, column=c).fill = alt_row_fill

            # =========================
            # VARIANCE COLOR
            # =========================

            variance = item.get("variance", 0)

            if variance > 0:
                variance_cell.fill = variance_positive

            elif variance < 0:
                variance_cell.fill = variance_negative

            row += 1

        # =========================
        # COLUMN WIDTHS
        # =========================

        widths = [
            28,  # Product
            14,  # Variant
            28,  # SKU
            18,  # Category
            16,  # Opening
            14,  # Added
            12,  # Adj+
            12,  # Adj-
            12,  # Sales
            18,  # Expected
            18,  # Actual
            14,  # Variance
            18,  # Value
        ]

        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(filepath)