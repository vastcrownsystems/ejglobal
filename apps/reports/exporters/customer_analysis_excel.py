from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class CustomerAnalysisExcelExporter:

    def export(self, data, filepath):

        wb = Workbook()
        ws = wb.active
        ws.title = "Customer Analysis"

        title_fill = PatternFill("solid", fgColor="0D47A1")
        header_fill = PatternFill("solid", fgColor="0D47A1")

        title_font = Font(color="FFFFFF", bold=True, size=16)
        header_font = Font(color="FFFFFF", bold=True)

        center = Alignment(horizontal="center")

        ws.merge_cells("A1:H1")

        title_cell = ws["A1"]
        title_cell.value = "CUSTOMER ANALYSIS REPORT"
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center

        ws.merge_cells("A2:H2")

        ws["A2"] = f"Period: {data['date_range']}"
        ws["A2"].alignment = center

        row = 4

        summary = data["summary"]

        ws["A4"] = "Total Customers"
        ws["B4"] = summary.get("total_customers", 0)

        ws["A5"] = "Total Revenue"
        ws["B5"] = summary.get("total_revenue", 0)
        ws["B5"].number_format = '"₦"#,##0.00'

        ws["A6"] = "Average Customer Value"
        ws["B6"] = summary.get("avg_customer_value", 0)
        ws["B6"].number_format = '"₦"#,##0.00'

        row = 9

        headers = [
            "Customer",
            "Phone",
            "Orders",
            "Items Bought",
            "Total Spent",
            "Avg Order",
            "First Purchase",
            "Last Purchase"
        ]

        for col, h in enumerate(headers,1):

            cell = ws.cell(row=row,column=col)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        row += 1

        for c in data["customers"]:

            ws.cell(row=row,column=1,value=c["customer"])
            ws.cell(row=row,column=2,value=c["phone"])
            ws.cell(row=row,column=3,value=c["orders"])
            ws.cell(row=row,column=4,value=c["items"])

            spent = ws.cell(row=row,column=5,value=c["revenue"])
            spent.number_format = '"₦"#,##0.00'

            avg = ws.cell(row=row,column=6,value=c["avg_order"])
            avg.number_format = '"₦"#,##0.00'

            ws.cell(row=row,column=7,value=str(c["first_purchase"])[:19])
            ws.cell(row=row,column=8,value=str(c["last_purchase"])[:19])

            row += 1

        widths = [26,18,10,14,16,16,20,20]

        for i,w in enumerate(widths,1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(filepath)