from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class OverdueCreditExcelExporter:

    def export(self,data,filepath):

        wb = Workbook()
        ws = wb.active
        ws.title = "Overdue Credit"

        title_fill = PatternFill("solid", fgColor="0D47A1")
        header_fill = PatternFill("solid", fgColor="0D47A1")

        title_font = Font(color="FFFFFF", bold=True, size=16)
        header_font = Font(color="FFFFFF", bold=True)

        center = Alignment(horizontal="center")

        ws.merge_cells("A1:H1")

        title = ws["A1"]
        title.value = "OVERDUE CREDIT REPORT"
        title.fill = title_fill
        title.font = title_font
        title.alignment = center

        ws.merge_cells("A2:H2")
        ws["A2"] = data["date_range"]
        ws["A2"].alignment = center

        headers = [
            "Customer","Phone","Invoice",
            "Due Date","Days Overdue",
            "Outstanding","Status"
        ]

        row = 4

        for col,h in enumerate(headers,1):

            cell = ws.cell(row=row,column=col)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        row += 1

        for r in data["rows"]:

            ws.cell(row=row,column=1,value=r["customer"])
            ws.cell(row=row,column=2,value=r["phone"])
            ws.cell(row=row,column=3,value=r["invoice"])
            ws.cell(row=row,column=4,value=str(r["due_date"]))
            ws.cell(row=row,column=5,value=r["days_overdue"])

            amt = ws.cell(row=row,column=6,value=r["outstanding"])
            amt.number_format = '"₦"#,##0.00'

            ws.cell(row=row,column=7,value=r["status"])

            row += 1

        widths=[26,18,16,14,14,16,14]

        for i,w in enumerate(widths,1):
            ws.column_dimensions[get_column_letter(i)].width=w

        wb.save(filepath)