from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class CashierPerformanceExcelExporter:

    @staticmethod
    def export(data, path):

        wb = Workbook()
        ws = wb.active
        ws.title = "Cashier Performance"

        header_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Title
        title_fill = PatternFill("solid", fgColor="0D47A1")
        title_font = Font(color="FFFFFF", bold=True, size=16)
        center = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A1:H1")

        title_cell = ws["A1"]
        title_cell.value = "CASHIER PERFORMANCE REPORT"
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center

        ws.merge_cells("A2:H2")
        ws["A2"] = f"Period: {data['date_range']}"
        ws["A2"].alignment = Alignment(horizontal="center")

        row = 4

        summary = data["summary"]

        summary_rows = [
            ("Total Cashiers", summary["total_cashiers"]),
            ("Total Orders", summary["total_orders"]),
            ("Total Revenue", f"₦{summary['total_revenue']:,.2f}"),
            ("Total Sessions", summary["total_sessions"]),
        ]

        for label, value in summary_rows:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = bold_font
            ws[f"B{row}"] = value
            row += 1

        row += 2

        # Cashier performance table
        headers = [
            "Cashier",
            "Orders",
            "Items Sold",
            "Revenue",
            "Avg Order",
            "First Sale",
            "Last Sale"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        row += 1

        for c in data["cashiers"]:

            values = [
                c["cashier"],
                c["orders"],
                c["items_sold"],
                f"₦{c['revenue']:,.2f}",
                f"₦{c['avg_order']:,.2f}",
                str(c["first_sale"])[:19],
                str(c["last_sale"])[:19]
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border

            row += 1

        row += 3

        # Session analysis
        ws[f"A{row}"] = "SESSION ANALYSIS"
        ws[f"A{row}"].font = Font(size=14, bold=True)

        row += 2

        session_headers = [
            "Session",
            "Store",
            "Register",
            "Cashier",
            "Opened",
            "Closed",
            "Orders",
            "Revenue"
        ]

        for col, header in enumerate(session_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        row += 1

        for s in data["sessions"]:

            values = [
                s["session"],
                s["store"],
                s["register"],
                s["cashier"],
                str(s["opened"])[:19],
                str(s["closed"])[:19] if s["closed"] else "",
                s["orders"],
                f"₦{s['revenue']:,.2f}"
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border

            row += 1

        # Column width
        widths = [22, 12, 14, 16, 16, 20, 20]

        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(path)