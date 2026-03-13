from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class SalesSummaryExcelExporter:

    def export(self, data, filepath):

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Summary"

        summary = data["summary"]

        # --------------------------------
        # Styles
        # --------------------------------

        header_fill = PatternFill("solid", fgColor="0D47A1")
        header_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left")

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        currency_format = '"₦"#,##0.00'

        # --------------------------------
        # Title
        # --------------------------------

        ws.merge_cells("A1:F1")
        ws["A1"] = "SALES SUMMARY REPORT"
        ws["A1"].fill = header_fill
        ws["A1"].font = header_font
        ws["A1"].alignment = center

        ws.merge_cells("A2:F2")
        ws["A2"] = f"Period: {data.get('date_range','')}"
        ws["A2"].alignment = center

        # --------------------------------
        # Summary
        # --------------------------------

        ws["A4"] = "Total Orders"
        ws["B4"] = summary["total_orders"]

        ws["A5"] = "Total Items Sold"
        ws["B5"] = summary["total_items"]

        ws["A6"] = "Total Revenue"
        ws["B6"] = summary["total_revenue"]
        ws["B6"].number_format = currency_format

        # --------------------------------
        # Payment Breakdown
        # --------------------------------

        ws["A8"] = "PAYMENT BREAKDOWN"
        ws["A8"].font = bold_font

        ws["A9"] = "Cash Sales"
        ws["B9"] = summary["cash_sales"]
        ws["B9"].number_format = currency_format

        ws["A10"] = "Card Sales"
        ws["B10"] = summary["card_sales"]
        ws["B10"].number_format = currency_format

        ws["A11"] = "Transfer Sales"
        ws["B11"] = summary["transfer_sales"]
        ws["B11"].number_format = currency_format

        ws["A12"] = "Credit Sales"
        ws["B12"] = summary["credit_sales"]
        ws["B12"].number_format = currency_format

        ws["A14"] = "Total Paid"
        ws["B14"] = summary["total_paid"]
        ws["B14"].number_format = currency_format

        ws["A15"] = "Credit Sold"
        ws["B15"] = summary["credit_sold"]
        ws["B15"].number_format = currency_format

        # --------------------------------
        # Product Sales Table
        # --------------------------------

        row = 18

        headers = [
            "Product",
            "Variant",
            "SKU",
            "Qty Sold",
            "Unit Price",
            "Revenue"
        ]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        row += 1

        for p in data["products"]:

            ws.cell(row=row, column=1).value = p["product_name"]
            ws.cell(row=row, column=2).value = p["variant_name"]
            ws.cell(row=row, column=3).value = p["sku"]
            ws.cell(row=row, column=4).value = p["quantity_sold"]

            price_cell = ws.cell(row=row, column=5)
            price_cell.value = p["unit_price"]
            price_cell.number_format = currency_format

            revenue_cell = ws.cell(row=row, column=6)
            revenue_cell.value = p["revenue"]
            revenue_cell.number_format = currency_format

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border

            row += 1

        # --------------------------------
        # Daily Sales Trend
        # --------------------------------

        row += 3

        ws.cell(row=row, column=1).value = "Daily Sales Trend"
        ws.cell(row=row, column=1).font = bold_font

        row += 1

        headers = ["Date", "Orders", "Revenue"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        row += 1

        for t in data.get("trend", []):

            ws.cell(row=row, column=1).value = str(t["date"])
            ws.cell(row=row, column=2).value = t["orders"]

            revenue = ws.cell(row=row, column=3)
            revenue.value = t["revenue"]
            revenue.number_format = currency_format

            for col in range(1, 4):
                ws.cell(row=row, column=col).border = border

            row += 1

        # --------------------------------
        # Column Widths
        # --------------------------------

        widths = [28, 12, 28, 12, 14, 16]

        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(filepath)